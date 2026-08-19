from datetime import date, timedelta
from io import BytesIO

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.admin.views.decorators import staff_member_required
from django.db import transaction
from django.db.models import Count, Q
from django.http import HttpResponse
from django.urls import reverse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
import qrcode
from qrcode.constants import ERROR_CORRECT_M

from .forms import (
    AbnormalIssueUpdateForm, EquipmentForm, InspectionAnswerFormSet, InspectionItemForm,
    InspectionRecordForm, InspectionTemplateForm,
)
from .models import AbnormalIssue, AbnormalIssueUpdate, Equipment, InspectionAnswer, InspectionItem, InspectionRecord, InspectionTemplate


@login_required
def dashboard(request):
    today = timezone.localdate()
    all_templates = (
        InspectionTemplate.objects.filter(is_active=True, equipment__is_active=True)
        .select_related('equipment')
        .prefetch_related('items')
    )
    templates = [template for template in all_templates if template.runs_on(today)]
    today_records = {
        record.template_id: record
        for record in InspectionRecord.objects.filter(
            inspection_date=today,
            template__in=templates,
        ).select_related('template', 'template__equipment', 'inspected_by')
    }

    cards = [{'template': template, 'record': today_records.get(template.id)} for template in templates]
    completed_count = len(today_records)
    abnormal_count = sum(1 for record in today_records.values() if record.status == InspectionRecord.Status.ABNORMAL)

    recent_abnormal_records = (
        InspectionRecord.objects.filter(status=InspectionRecord.Status.ABNORMAL)
        .select_related('template', 'template__equipment', 'inspected_by')
        .order_by('-inspection_date', '-updated_at')[:5]
    )
    unresolved_issues = (
        AbnormalIssue.objects.exclude(status=AbnormalIssue.Status.RESOLVED)
        .select_related('answer__record__template__equipment', 'answer__item')
        .prefetch_related('updates__updated_by')
        .order_by('answer__record__inspection_date', 'created_at')
    )

    return render(request, 'inspections/dashboard.html', {
        'today': today,
        'cards': cards,
        'completed_count': completed_count,
        'pending_count': max(len(templates) - completed_count, 0),
        'abnormal_count': abnormal_count,
        'recent_abnormal_records': recent_abnormal_records,
        'unresolved_issues': unresolved_issues[:8],
        'unresolved_count': unresolved_issues.count(),
    })


@login_required
@transaction.atomic
def perform_inspection(request, template_id):
    template = get_object_or_404(
        InspectionTemplate.objects.select_related('equipment').prefetch_related('items'),
        pk=template_id,
        is_active=True,
        equipment__is_active=True,
    )
    today = timezone.localdate()
    if not template.runs_on(today):
        messages.warning(request, f'この点検表は本日（{template.WEEKDAY_LABELS[today.weekday()]}曜日）の実施対象ではありません。')
        return redirect('dashboard')

    existing = InspectionRecord.objects.filter(template=template, inspection_date=today).first()
    initial_answers = []
    existing_answers = {}
    if existing:
        existing_answers = {answer.item_id: answer for answer in existing.answers.select_related('item')}

    for item in template.items.all():
        answer = existing_answers.get(item.id)
        initial_answers.append({
            'item_id': item.id,
            'label': item.label,
            'required_item': item.is_required,
            'result': answer.result if answer else '',
            'note': answer.note if answer else '',
            'existing_photo_url': answer.photo.url if answer and answer.photo else '',
        })

    if request.method == 'POST':
        record_form = InspectionRecordForm(request.POST, instance=existing)
        answer_formset = InspectionAnswerFormSet(request.POST, request.FILES, initial=initial_answers)

        if record_form.is_valid() and answer_formset.is_valid():
            has_answer_error = False
            any_abnormal = False

            for form in answer_formset:
                result = form.cleaned_data.get('result', '')
                required_item = form.cleaned_data.get('required_item', False)
                note = (form.cleaned_data.get('note') or '').strip()

                if required_item and not result:
                    form.add_error('result', '必須項目です。「正常」または「異常」を選択してください。')
                    has_answer_error = True

                if result == InspectionAnswer.Result.ABNORMAL:
                    any_abnormal = True
                    if not note:
                        form.add_error('note', '「異常」を選択した場合は内容を入力してください。')
                        has_answer_error = True

            if not has_answer_error:
                record = record_form.save(commit=False)
                record.template = template
                record.inspection_date = today
                record.inspected_by = request.user
                record.status = (
                    InspectionRecord.Status.ABNORMAL
                    if any_abnormal
                    else InspectionRecord.Status.NORMAL
                )
                record.save()

                valid_item_ids = {item.id for item in template.items.all()}
                for form in answer_formset:
                    item_id = form.cleaned_data['item_id']
                    if item_id not in valid_item_ids:
                        continue

                    result = form.cleaned_data.get('result', '')
                    uploaded_photo = form.cleaned_data.get('photo')
                    remove_photo = form.cleaned_data.get('remove_photo', False)

                    answer, _ = InspectionAnswer.objects.get_or_create(
                        record=record,
                        item_id=item_id,
                    )
                    old_photo = answer.photo if answer.photo else None

                    answer.result = result
                    answer.checked = bool(result)
                    answer.note = (form.cleaned_data.get('note') or '').strip()

                    # 写真は異常項目だけに保持する。正常へ戻した場合は削除する。
                    if result != InspectionAnswer.Result.ABNORMAL:
                        if old_photo:
                            old_photo.delete(save=False)
                        answer.photo = None
                    elif remove_photo:
                        if old_photo:
                            old_photo.delete(save=False)
                        answer.photo = None
                    elif uploaded_photo:
                        if old_photo:
                            old_photo.delete(save=False)
                        answer.photo = uploaded_photo

                    answer.save()

                    # 異常項目は解決まで追跡する。誤入力修正で正常に戻した場合は自動的に対応済みにする。
                    if result == InspectionAnswer.Result.ABNORMAL:
                        AbnormalIssue.objects.get_or_create(answer=answer)
                    else:
                        issue = AbnormalIssue.objects.filter(answer=answer).first()
                        if issue and issue.status != AbnormalIssue.Status.RESOLVED:
                            issue.status = AbnormalIssue.Status.RESOLVED
                            issue.resolved_at = timezone.now()
                            issue.save(update_fields=['status', 'resolved_at', 'updated_at'])
                            AbnormalIssueUpdate.objects.create(
                                issue=issue, status=AbnormalIssue.Status.RESOLVED,
                                note='点検結果が正常へ修正されたため、自動的に対応済みに変更しました。',
                                updated_by=request.user,
                            )

                if any_abnormal:
                    messages.warning(request, f'{template.equipment.name}の{template.name}を「異常あり」で保存しました。')
                else:
                    messages.success(request, f'{template.equipment.name}の{template.name}を保存しました。')
                return redirect('dashboard')
    else:
        record_form = InspectionRecordForm(instance=existing)
        answer_formset = InspectionAnswerFormSet(initial=initial_answers)

    return render(request, 'inspections/perform_inspection.html', {
        'template': template,
        'today': today,
        'record_form': record_form,
        'answer_formset': answer_formset,
        'existing': existing,
    })


@login_required
def equipment_scan(request, equipment_id):
    """QRコードから装置の本日点検へ案内する。

    点検表が1つなら直接点検画面へ、複数なら選択画面へ遷移する。
    """
    equipment = get_object_or_404(Equipment, pk=equipment_id, is_active=True)
    today = timezone.localdate()
    active_templates = list(
        equipment.templates.filter(is_active=True).prefetch_related('items').order_by('name')
    )
    today_templates = [template for template in active_templates if template.runs_on(today)]

    if len(today_templates) == 1:
        return redirect('perform_inspection', template_id=today_templates[0].id)

    today_records = {
        record.template_id: record
        for record in InspectionRecord.objects.filter(
            inspection_date=today,
            template__in=today_templates,
        ).select_related('inspected_by')
    }
    rows = [
        {'template': template, 'record': today_records.get(template.id)}
        for template in today_templates
    ]

    return render(request, 'inspections/equipment_scan.html', {
        'equipment': equipment,
        'today': today,
        'rows': rows,
        'active_templates': active_templates,
    })


@staff_member_required(login_url='login')
def equipment_qr_page(request, equipment_id):
    equipment = get_object_or_404(Equipment, pk=equipment_id)
    scan_url = request.build_absolute_uri(
        reverse('equipment_scan', args=[equipment.id])
    )
    return render(request, 'inspections/equipment_qr.html', {
        'equipment': equipment,
        'scan_url': scan_url,
    })


@staff_member_required(login_url='login')
def equipment_qr_image(request, equipment_id):
    equipment = get_object_or_404(Equipment, pk=equipment_id)
    scan_url = request.build_absolute_uri(
        reverse('equipment_scan', args=[equipment.id])
    )

    qr = qrcode.QRCode(
        version=None,
        error_correction=ERROR_CORRECT_M,
        box_size=9,
        border=4,
    )
    qr.add_data(scan_url)
    qr.make(fit=True)
    image = qr.make_image(fill_color='black', back_color='white')

    output = BytesIO()
    image.save(output, format='PNG')
    response = HttpResponse(output.getvalue(), content_type='image/png')
    response['Cache-Control'] = 'no-store, max-age=0'
    response['Content-Disposition'] = f'inline; filename=equipment_{equipment.id}_qr.png'
    return response


def _parse_iso_date(value, fallback):
    try:
        return date.fromisoformat(value) if value else fallback
    except (TypeError, ValueError):
        return fallback


def _iter_dates(start_date, end_date):
    current = start_date
    while current <= end_date:
        yield current
        current += timedelta(days=1)


def _iter_month_starts(start_date, end_date):
    current = start_date.replace(day=1)
    last = end_date.replace(day=1)
    while current <= last:
        yield current
        if current.month == 12:
            current = current.replace(year=current.year + 1, month=1)
        else:
            current = current.replace(month=current.month + 1)


def _analytics_data(request):
    """集計画面とExcel出力で共通利用する集計データを作る。"""
    today = timezone.localdate()
    default_start = today.replace(day=1)
    start_date = _parse_iso_date(request.GET.get('start', '').strip(), default_start)
    end_date = _parse_iso_date(request.GET.get('end', '').strip(), today)
    if start_date > end_date:
        start_date, end_date = end_date, start_date

    equipment_id = request.GET.get('equipment', '').strip()
    equipment_choices = Equipment.objects.order_by('display_order', 'name')
    selected_equipment = None
    if equipment_id.isdigit():
        selected_equipment = Equipment.objects.filter(pk=int(equipment_id)).first()

    records = InspectionRecord.objects.filter(
        inspection_date__range=(start_date, end_date)
    ).select_related('template', 'template__equipment', 'inspected_by')
    answers = InspectionAnswer.objects.filter(
        record__inspection_date__range=(start_date, end_date),
        result=InspectionAnswer.Result.ABNORMAL,
    ).select_related('record', 'record__template', 'record__template__equipment', 'item')

    templates = InspectionTemplate.objects.filter(
        is_active=True,
        equipment__is_active=True,
    ).select_related('equipment')

    if selected_equipment:
        records = records.filter(template__equipment=selected_equipment)
        answers = answers.filter(record__template__equipment=selected_equipment)
        templates = templates.filter(equipment=selected_equipment)

    records_list = list(records.order_by('inspection_date', 'template__equipment__display_order', 'template__name'))
    templates_list = list(templates)

    # 現在の点検表設定を基準に、指定期間で実施対象だった回数を算出する。
    expected_by_equipment = {}
    expected_total = 0
    date_list = list(_iter_dates(start_date, end_date))
    for template in templates_list:
        expected = sum(1 for target_date in date_list if template.runs_on(target_date))
        expected_total += expected
        expected_by_equipment[template.equipment_id] = expected_by_equipment.get(template.equipment_id, 0) + expected

    completed_total = len(records_list)
    abnormal_total = sum(1 for record in records_list if record.status == InspectionRecord.Status.ABNORMAL)
    unresolved_query = AbnormalIssue.objects.exclude(status=AbnormalIssue.Status.RESOLVED).filter(answer__record__inspection_date__lte=end_date)
    if selected_equipment:
        unresolved_query = unresolved_query.filter(answer__record__template__equipment=selected_equipment)
    unresolved_total = unresolved_query.count()
    pending_total = max(expected_total - completed_total, 0)
    implementation_rate = round((completed_total / expected_total) * 100, 1) if expected_total else 0.0

    completed_group = {
        row['template__equipment_id']: row['count']
        for row in records.values('template__equipment_id').annotate(count=Count('id'))
    }
    abnormal_group = {
        row['template__equipment_id']: row['count']
        for row in records.filter(status=InspectionRecord.Status.ABNORMAL)
        .values('template__equipment_id').annotate(count=Count('id'))
    }

    if selected_equipment:
        equipment_for_stats = [selected_equipment]
    else:
        equipment_for_stats = list(Equipment.objects.filter(is_active=True).order_by('display_order', 'name'))

    equipment_stats = []
    for equipment in equipment_for_stats:
        expected = expected_by_equipment.get(equipment.id, 0)
        completed = completed_group.get(equipment.id, 0)
        abnormal = abnormal_group.get(equipment.id, 0)
        rate = round((completed / expected) * 100, 1) if expected else 0.0
        equipment_stats.append({
            'equipment': equipment,
            'expected': expected,
            'completed': completed,
            'pending': max(expected - completed, 0),
            'abnormal': abnormal,
            'rate': rate,
            'rate_bar': min(rate, 100),
        })

    item_rows = list(
        answers.values('item_id', 'item__label', 'record__template__equipment__name')
        .annotate(count=Count('id'))
        .order_by('-count', 'record__template__equipment__name', 'item__label')[:15]
    )
    max_item_count = max((row['count'] for row in item_rows), default=0)
    item_ranking = [
        {
            'label': row['item__label'],
            'equipment_name': row['record__template__equipment__name'],
            'count': row['count'],
            'bar': round((row['count'] / max_item_count) * 100, 1) if max_item_count else 0,
        }
        for row in item_rows
    ]

    monthly_trend = []
    for month_start in _iter_month_starts(start_date, end_date):
        if month_start.month == 12:
            next_month = month_start.replace(year=month_start.year + 1, month=1)
        else:
            next_month = month_start.replace(month=month_start.month + 1)
        month_end = min(next_month - timedelta(days=1), end_date)
        month_begin = max(month_start, start_date)
        month_records = [r for r in records_list if month_begin <= r.inspection_date <= month_end]
        abnormal_count = sum(1 for r in month_records if r.status == InspectionRecord.Status.ABNORMAL)
        monthly_trend.append({
            'label': f'{month_start.year}/{month_start.month}',
            'completed': len(month_records),
            'abnormal': abnormal_count,
        })
    max_month_abnormal = max((row['abnormal'] for row in monthly_trend), default=0)
    for row in monthly_trend:
        row['bar'] = round((row['abnormal'] / max_month_abnormal) * 100, 1) if max_month_abnormal else 0

    return {
        'start_date': start_date,
        'end_date': end_date,
        'selected_equipment': selected_equipment,
        'equipment_choices': equipment_choices,
        'records': records_list,
        'expected_total': expected_total,
        'completed_total': completed_total,
        'pending_total': pending_total,
        'abnormal_total': abnormal_total,
        'unresolved_total': unresolved_total,
        'implementation_rate': implementation_rate,
        'implementation_bar': min(implementation_rate, 100),
        'equipment_stats': equipment_stats,
        'item_ranking': item_ranking,
        'monthly_trend': monthly_trend,
    }


@login_required
def analytics(request):
    data = _analytics_data(request)
    return render(request, 'inspections/analytics.html', data)


@login_required
def export_analytics_excel(request):
    data = _analytics_data(request)
    workbook = Workbook()
    summary = workbook.active
    summary.title = '集計サマリー'

    header_fill = PatternFill('solid', fgColor='D9EAF7')
    accent_fill = PatternFill('solid', fgColor='EAF6EF')

    summary.append(['集計期間', f"{data['start_date']:%Y-%m-%d} ～ {data['end_date']:%Y-%m-%d}"])
    summary.append(['装置', data['selected_equipment'].name if data['selected_equipment'] else 'すべて'])
    summary.append([])
    summary.append(['指標', '値'])
    for cell in summary[4]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    summary.append(['実施対象', data['expected_total']])
    summary.append(['実施済み', data['completed_total']])
    summary.append(['未実施', data['pending_total']])
    summary.append(['異常あり', data['abnormal_total']])
    summary.append(['未解決異常', data['unresolved_total']])
    summary.append(['点検実施率', f"{data['implementation_rate']}%"])
    summary.column_dimensions['A'].width = 20
    summary.column_dimensions['B'].width = 28

    equipment_sheet = workbook.create_sheet('装置別集計')
    equipment_sheet.append(['装置', '実施対象', '実施済み', '未実施', '実施率', '異常件数'])
    for cell in equipment_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in data['equipment_stats']:
        equipment_sheet.append([
            row['equipment'].name, row['expected'], row['completed'], row['pending'],
            f"{row['rate']}%", row['abnormal'],
        ])
    for idx, width in enumerate([22, 12, 12, 12, 12, 12], start=1):
        equipment_sheet.column_dimensions[get_column_letter(idx)].width = width
    equipment_sheet.freeze_panes = 'A2'

    item_sheet = workbook.create_sheet('異常項目ランキング')
    item_sheet.append(['順位', '装置', '点検項目', '異常回数'])
    for cell in item_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for rank, row in enumerate(data['item_ranking'], start=1):
        item_sheet.append([rank, row['equipment_name'], row['label'], row['count']])
    for idx, width in enumerate([8, 22, 42, 12], start=1):
        item_sheet.column_dimensions[get_column_letter(idx)].width = width

    monthly_sheet = workbook.create_sheet('月別推移')
    monthly_sheet.append(['月', '実施件数', '異常件数'])
    for cell in monthly_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in data['monthly_trend']:
        monthly_sheet.append([row['label'], row['completed'], row['abnormal']])
    for idx, width in enumerate([14, 14, 14], start=1):
        monthly_sheet.column_dimensions[get_column_letter(idx)].width = width

    records_sheet = workbook.create_sheet('対象点検履歴')
    records_sheet.append(['点検日', '装置', '点検表', '点検者', '結果'])
    for cell in records_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = accent_fill
    for record in data['records']:
        records_sheet.append([
            record.inspection_date.strftime('%Y-%m-%d'),
            record.template.equipment.name,
            record.template.name,
            record.inspected_by.get_full_name() or record.inspected_by.username,
            record.get_status_display(),
        ])
    for idx, width in enumerate([12, 22, 22, 18, 12], start=1):
        records_sheet.column_dimensions[get_column_letter(idx)].width = width
    records_sheet.freeze_panes = 'A2'

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"inspection_analytics_{data['start_date']:%Y%m%d}_{data['end_date']:%Y%m%d}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def history(request):
    records = InspectionRecord.objects.select_related('template', 'template__equipment', 'inspected_by')
    keyword = request.GET.get('q', '').strip()
    status = request.GET.get('status', '').strip()
    start = request.GET.get('start', '').strip()
    end = request.GET.get('end', '').strip()

    if keyword:
        records = records.filter(
            Q(template__equipment__name__icontains=keyword)
            | Q(template__name__icontains=keyword)
            | Q(abnormal_details__icontains=keyword)
            | Q(action_taken__icontains=keyword)
            | Q(inspected_by__username__icontains=keyword)
        )
    if status in {InspectionRecord.Status.NORMAL, InspectionRecord.Status.ABNORMAL}:
        records = records.filter(status=status)
    if start:
        records = records.filter(inspection_date__gte=start)
    if end:
        records = records.filter(inspection_date__lte=end)

    return render(request, 'inspections/history.html', {
        'records': records[:300],
        'filters': {'q': keyword, 'status': status, 'start': start, 'end': end},
    })


@login_required
def record_detail(request, record_id):
    record = get_object_or_404(
        InspectionRecord.objects.select_related('template', 'template__equipment', 'inspected_by').prefetch_related('answers__item', 'answers__issue__updates__updated_by'),
        pk=record_id,
    )
    return render(request, 'inspections/record_detail.html', {'record': record})


@login_required
def issues(request):
    status = request.GET.get('status', 'unresolved').strip()
    queryset = (
        AbnormalIssue.objects.select_related(
            'answer__record__template__equipment', 'answer__record__inspected_by', 'answer__item'
        ).prefetch_related('updates__updated_by')
    )
    if status == 'unresolved':
        queryset = queryset.exclude(status=AbnormalIssue.Status.RESOLVED)
    elif status in {value for value, _ in AbnormalIssue.Status.choices}:
        queryset = queryset.filter(status=status)
    return render(request, 'inspections/issues.html', {
        'issues': queryset.order_by('-answer__record__inspection_date', '-updated_at')[:300],
        'status_filter': status,
    })


@login_required
@transaction.atomic
def issue_detail(request, issue_id):
    issue = get_object_or_404(
        AbnormalIssue.objects.select_related(
            'answer__record__template__equipment', 'answer__record__inspected_by', 'answer__item'
        ).prefetch_related('updates__updated_by'),
        pk=issue_id,
    )
    if request.method == 'POST':
        form = AbnormalIssueUpdateForm(request.POST)
        if form.is_valid():
            new_status = form.cleaned_data['status']
            AbnormalIssueUpdate.objects.create(
                issue=issue, status=new_status, note=form.cleaned_data['note'].strip(), updated_by=request.user
            )
            issue.status = new_status
            issue.resolved_at = timezone.now() if new_status == AbnormalIssue.Status.RESOLVED else None
            issue.save(update_fields=['status', 'resolved_at', 'updated_at'])
            messages.success(request, '対応内容を記録しました。')
            return redirect('issue_detail', issue_id=issue.id)
    else:
        form = AbnormalIssueUpdateForm(initial={'status': issue.status})
    return render(request, 'inspections/issue_detail.html', {'issue': issue, 'form': form})


@login_required
def export_excel(request):
    records = InspectionRecord.objects.select_related(
        'template', 'template__equipment', 'inspected_by'
    ).prefetch_related('answers__item')

    start = request.GET.get('start', '').strip()
    end = request.GET.get('end', '').strip()
    if start:
        records = records.filter(inspection_date__gte=start)
    if end:
        records = records.filter(inspection_date__lte=end)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '点検履歴'
    headers = ['点検日', '装置・部屋', '点検表', '点検者', '結果', '異常内容', '対応内容', '登録日時']
    sheet.append(headers)

    header_fill = PatternFill('solid', fgColor='D9EAF7')
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for record in records:
        sheet.append([
            record.inspection_date.strftime('%Y-%m-%d'),
            record.template.equipment.name,
            record.template.name,
            record.inspected_by.get_full_name() or record.inspected_by.username,
            record.get_status_display(),
            record.abnormal_details,
            record.action_taken,
            timezone.localtime(record.created_at).strftime('%Y-%m-%d %H:%M'),
        ])

    widths = [12, 20, 18, 16, 12, 35, 35, 18]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions

    detail_sheet = workbook.create_sheet('点検項目')
    detail_headers = ['点検日', '装置・部屋', '点検表', '点検項目', '結果', '異常内容・メモ', '写真']
    detail_sheet.append(detail_headers)
    for cell in detail_sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for record in records:
        for answer in record.answers.all():
            detail_sheet.append([
                record.inspection_date.strftime('%Y-%m-%d'),
                record.template.equipment.name,
                record.template.name,
                answer.item.label,
                answer.get_result_display() if answer.result else '未入力',
                answer.note,
                'あり' if answer.photo else '',
            ])
    for index, width in enumerate([12, 20, 18, 40, 10, 30, 10], start=1):
        detail_sheet.column_dimensions[get_column_letter(index)].width = width
    detail_sheet.freeze_panes = 'A2'
    detail_sheet.auto_filter.ref = detail_sheet.dimensions

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f'inspection_history_{date.today().isoformat()}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@staff_member_required(login_url='login')
def settings_home(request):
    equipment_list = Equipment.objects.prefetch_related('templates__items').all()
    return render(request, 'inspections/settings_home.html', {
        'equipment_list': equipment_list,
    })


@staff_member_required(login_url='login')
def equipment_create(request):
    form = EquipmentForm(request.POST or None)
    if request.method == 'POST' and form.is_valid():
        equipment = form.save()
        messages.success(request, f'「{equipment.name}」を追加しました。')
        return redirect('settings_home')
    return render(request, 'inspections/settings_form.html', {
        'form': form,
        'title': '装置・部屋を追加',
        'submit_label': '追加する',
    })


@staff_member_required(login_url='login')
def equipment_edit(request, equipment_id):
    equipment = get_object_or_404(Equipment, pk=equipment_id)
    form = EquipmentForm(request.POST or None, instance=equipment)
    if request.method == 'POST' and form.is_valid():
        equipment = form.save()
        messages.success(request, f'「{equipment.name}」を更新しました。')
        return redirect('settings_home')
    return render(request, 'inspections/settings_form.html', {
        'form': form,
        'title': '装置・部屋を編集',
        'submit_label': '更新する',
        'object': equipment,
    })


@staff_member_required(login_url='login')
def equipment_delete(request, equipment_id):
    equipment = get_object_or_404(Equipment, pk=equipment_id)
    if request.method == 'POST':
        if InspectionRecord.objects.filter(template__equipment=equipment).exists():
            messages.error(request, '点検履歴がある装置は削除できません。「使用中」のチェックを外してください。')
        else:
            name = equipment.name
            equipment.delete()
            messages.success(request, f'「{name}」を削除しました。')
        return redirect('settings_home')
    return render(request, 'inspections/settings_confirm_delete.html', {
        'title': '装置・部屋を削除',
        'object': equipment,
        'warning': 'この装置に含まれる点検表と点検項目も削除されます。',
    })


@staff_member_required(login_url='login')
def template_create(request, equipment_id=None):
    initial = {}
    if equipment_id:
        initial['equipment'] = get_object_or_404(Equipment, pk=equipment_id)
    form = InspectionTemplateForm(request.POST or None, initial=initial)
    if request.method == 'POST' and form.is_valid():
        template = form.save()
        messages.success(request, f'「{template}」を追加しました。')
        return redirect('settings_home')
    return render(request, 'inspections/settings_form.html', {
        'form': form,
        'title': '点検表を追加',
        'submit_label': '追加する',
    })


@staff_member_required(login_url='login')
def template_edit(request, template_id):
    template = get_object_or_404(InspectionTemplate, pk=template_id)
    form = InspectionTemplateForm(request.POST or None, instance=template)
    if request.method == 'POST' and form.is_valid():
        template = form.save()
        messages.success(request, f'「{template}」を更新しました。')
        return redirect('settings_home')
    return render(request, 'inspections/settings_form.html', {
        'form': form,
        'title': '点検表を編集',
        'submit_label': '更新する',
        'object': template,
    })


@staff_member_required(login_url='login')
def template_delete(request, template_id):
    template = get_object_or_404(InspectionTemplate, pk=template_id)
    if request.method == 'POST':
        if template.records.exists():
            messages.error(request, '点検履歴がある点検表は削除できません。「使用中」のチェックを外してください。')
        else:
            name = str(template)
            template.delete()
            messages.success(request, f'「{name}」を削除しました。')
        return redirect('settings_home')
    return render(request, 'inspections/settings_confirm_delete.html', {
        'title': '点検表を削除',
        'object': template,
        'warning': 'この点検表に含まれる点検項目も削除されます。',
    })


@staff_member_required(login_url='login')
def item_create(request, template_id):
    template = get_object_or_404(InspectionTemplate, pk=template_id)
    next_order = (template.items.order_by('-display_order').values_list('display_order', flat=True).first() or 0) + 10
    form = InspectionItemForm(request.POST or None, initial={'display_order': next_order, 'is_required': True})
    if request.method == 'POST' and form.is_valid():
        item = form.save(commit=False)
        item.template = template
        item.save()
        messages.success(request, f'「{item.label}」を追加しました。')
        return redirect('settings_home')
    return render(request, 'inspections/settings_form.html', {
        'form': form,
        'title': f'{template}：点検項目を追加',
        'submit_label': '追加する',
    })


@staff_member_required(login_url='login')
def item_edit(request, item_id):
    item = get_object_or_404(InspectionItem.objects.select_related('template', 'template__equipment'), pk=item_id)
    form = InspectionItemForm(request.POST or None, instance=item)
    if request.method == 'POST' and form.is_valid():
        item = form.save()
        messages.success(request, f'「{item.label}」を更新しました。')
        return redirect('settings_home')
    return render(request, 'inspections/settings_form.html', {
        'form': form,
        'title': '点検項目を編集',
        'submit_label': '更新する',
        'object': item,
    })


@staff_member_required(login_url='login')
def item_delete(request, item_id):
    item = get_object_or_404(InspectionItem, pk=item_id)
    if request.method == 'POST':
        if item.answers.exists():
            messages.error(request, '過去の点検履歴で使われている項目は削除できません。項目名を「使用停止」などに変更してください。')
        else:
            label = item.label
            item.delete()
            messages.success(request, f'「{label}」を削除しました。')
        return redirect('settings_home')
    return render(request, 'inspections/settings_confirm_delete.html', {
        'title': '点検項目を削除',
        'object': item,
        'warning': '削除すると元に戻せません。',
    })
